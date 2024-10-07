from flask import Flask, render_template, redirect, url_for, request, flash, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
import csv
import os
import openpyxl
from io import BytesIO
import pdfkit 
from flask import jsonify


from openpyxl import load_workbook
from flask_mail import Mail, Message
from io import BytesIO

from models import db, User, Producto, Movimiento, Material, Herramienta, MovimientoMaterial, MovimientoHerramienta

app = Flask(__name__)
app.config['SECRET_KEY'] = 'tu_clave_secreta'  
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///db.sqlite3'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False  # Evitar advertencias

db.init_app(app)

login_manager = LoginManager()
login_manager.login_view = 'login'
login_manager.init_app(app)

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# Crear la base de datos con los modelos definidos
with app.app_context():
    db.create_all()
    
    # Verificar si ya existe un usuario administrador
    if not User.query.filter_by(role='admin').first():
        admin_user = User(username='admin', password=generate_password_hash('admin123', method='pbkdf2:sha256'), role='admin')
        db.session.add(admin_user)
        db.session.commit()
        print('Usuario administrador creado: admin / admin123')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password, password):
            login_user(user)
            if user.role == 'admin':
                return redirect(url_for('admin_dashboard'))
            else:
                return redirect(url_for('user_dashboard'))
        flash('Nombre de usuario o contraseña incorrectos')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/')
def home():
    return redirect(url_for('login'))

@app.route('/admin')
@login_required
def admin_dashboard():
    if current_user.role != 'admin':
        flash('Acceso denegado')
        return redirect(url_for('user_dashboard'))
    return render_template('admin_dashboard.html')

@app.route('/user')
@login_required
def user_dashboard():
    productos = Producto.query.all()
    return render_template('user_dashboard.html', productos=productos)

@app.route('/admin/producto/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo_producto():
    if current_user.role != 'admin':
        flash('Acceso denegado')
        return redirect(url_for('user_dashboard'))
    if request.method == 'POST':
        nombre = request.form['nombre']
        vida_util = request.form['vida_util']
        unidad = request.form['unidad']
        cantidad = request.form['cantidad']
        descripcion = request.form['descripcion']
        nuevo_producto = Producto(nombre=nombre, vida_util=vida_util, unidad=unidad, cantidad=cantidad, descripcion=descripcion)
        db.session.add(nuevo_producto)
        db.session.commit()
        return redirect(url_for('admin_dashboard'))
    return render_template('nuevo_producto.html')

@app.route('/admin/producto/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_producto(id):
    if current_user.role != 'admin':
        flash('Acceso denegado')
        return redirect(url_for('user_dashboard'))
    producto = Producto.query.get_or_404(id)
    if request.method == 'POST':
        producto.nombre = request.form['nombre']
        producto.vida_util = request.form['vida_util']
        producto.unidad = request.form['unidad']
        producto.cantidad = request.form['cantidad']
        producto.descripcion = request.form['descripcion']
        db.session.commit()
        return redirect(url_for('admin_dashboard'))
    return render_template('editar_producto.html', producto=producto)

@app.route('/admin/producto/eliminar/<int:id>')
@login_required
def eliminar_producto(id):
    if current_user.role != 'admin':
        flash('Acceso denegado')
        return redirect(url_for('user_dashboard'))
    producto = Producto.query.get_or_404(id)
    db.session.delete(producto)
    db.session.commit()
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/movimiento/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo_movimiento():
    if current_user.role != 'admin':
        flash('Acceso denegado')
        return redirect(url_for('user_dashboard'))
    
    if request.method == 'POST':
        producto_id = request.form['producto_id']
        tipo = request.form['tipo']
        cantidad = request.form['cantidad']
        trabajador = request.form.get('trabajador') if tipo == 'salida' else None
        
        movimiento = Movimiento(producto_id=producto_id, tipo=tipo, cantidad=cantidad, trabajador=trabajador)
        
        producto = Producto.query.get(producto_id)
        if tipo == 'entrada':
            producto.cantidad += int(cantidad)
        elif tipo == 'salida':
            producto.cantidad -= int(cantidad)
        
        db.session.add(movimiento)
        db.session.commit()
        return redirect(url_for('admin_dashboard'))
    
    productos = Producto.query.all()
    return render_template('nuevo_movimiento.html', productos=productos)

@app.route('/register', methods=['GET', 'POST'])
@login_required
def register():
    if current_user.role != 'admin':
        flash('Acceso denegado')
        return redirect(url_for('user_dashboard'))
    
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        role = request.form['role']  # 'admin' o 'user'
        
        hashed_password = generate_password_hash(password, method='pbkdf2:sha256')
        
        new_user = User(username=username, password=hashed_password, role=role)
        db.session.add(new_user)
        db.session.commit()
        
        flash('Usuario registrado exitosamente')
        return redirect(url_for('admin_dashboard'))
    
    return render_template('register.html')

@app.route('/buscar', methods=['GET', 'POST'])
@login_required
def buscar():
    if request.method == 'POST':
        trabajador = request.form['trabajador']
        movimientos = Movimiento.query.filter_by(trabajador=trabajador, tipo='salida').all()
        
        return render_template('resultado_busqueda.html', movimientos=movimientos, trabajador=trabajador)
    
    return render_template('buscar.html')

@app.route('/buscar_material', methods=['GET', 'POST'])
@login_required
def buscar_material():
    if request.method == 'GET':
        persona = request.args.get('persona')  # Capturar el parámetro 'persona' de la búsqueda
        if persona:
            materiales = Material.query.filter_by(persona_asignada=persona).all()
        else:
            materiales = Material.query.all()  # Cargar todos los materiales si no hay filtro
        return render_template('material_dashboard.html', materiales=materiales)



@app.route('/buscar_herramienta', methods=['GET', 'POST'])
@login_required
def buscar_herramienta():
    if request.method == 'POST':
        trabajador = request.form['trabajador']
        movimientos = MovimientoHerramienta.query.filter_by(trabajador=trabajador, tipo='salida').all()
        
        return render_template('resultado_busqueda_herramienta.html', movimientos=movimientos, trabajador=trabajador)
    
    return render_template('buscar_herramienta.html')

# Nuevas rutas para Material
@app.route('/material')
@login_required
def material_dashboard():
    materiales = Material.query.all()
    return render_template('material_dashboard.html', materiales=materiales)

@app.route('/material/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo_material():
    if current_user.role != 'admin':
        flash('Acceso denegado')
        return redirect(url_for('admin_dashboard'))
    if request.method == 'POST':
        nombre = request.form['nombre']
        unidad = request.form['unidad']
        cantidad = request.form['cantidad']
        proyecto = request.form['proyecto']  # Se captura el nuevo campo 'proyecto'
        area_colocacion = request.form['area_colocacion']  # Se captura el nuevo campo 'area_colocacion'

        nuevo_material = Material(
            nombre=nombre,
            unidad=unidad,
            cantidad=cantidad,
            proyecto=proyecto,  # Guardar proyecto
            area_colocacion=area_colocacion  # Guardar área de colocación
        )
        db.session.add(nuevo_material)
        db.session.commit()
        flash('Material agregado correctamente')
        return redirect(url_for('material_dashboard'))

    return render_template('nuevo_material.html')



@app.route('/material/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_material(id):
    if current_user.role != 'admin':
        flash('Acceso denegado')
        return redirect(url_for('material_dashboard'))
    material = Material.query.get_or_404(id)
    if request.method == 'POST':
        material.nombre = request.form['nombre']
        material.unidad = request.form['unidad']
        material.cantidad = request.form['cantidad']
        material.proyecto = request.form['proyecto']  # Cambiado a 'proyecto'
        material.area_colocacion = request.form['area_colocacion']  # Nuevo campo 'area_colocacion'   
        db.session.commit()
        return redirect(url_for('material_dashboard'))
    return render_template('editar_material.html', material=material)


@app.route('/material/eliminar/<int:id>')
@login_required
def eliminar_material(id):
    if current_user.role != 'admin':
        flash('Acceso denegado')
        return redirect(url_for('material_dashboard'))
    material = Material.query.get_or_404(id)
    db.session.delete(material)
    db.session.commit()
    return redirect(url_for('material_dashboard'))

@app.route('/material/movimiento/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo_movimiento_material():
    if current_user.role != 'admin':
        flash('Acceso denegado')
        return redirect(url_for('material_dashboard'))

    if request.method == 'POST':
        material_id = request.form['material_id']
        tipo = request.form['tipo']
        cantidad = int(request.form['cantidad'])
        trabajador = request.form.get('trabajador') if tipo == 'salida' else None

        material = Material.query.get(material_id)

        # Validar que la cantidad sea mayor que 0
        if cantidad <= 0:
            flash('La cantidad debe ser mayor a 0.')
            return redirect(url_for('nuevo_movimiento_material'))

        # Validar que no se pueda registrar una salida mayor a lo disponible en stock
        if tipo == 'salida' and cantidad > material.cantidad:
            flash(f'No hay suficiente stock para registrar esta salida. El stock disponible es {material.cantidad}.')
            return redirect(url_for('nuevo_movimiento_material'))

        # Si es entrada o salida, ajustar el stock
        if tipo == 'entrada':
            material.cantidad += cantidad
        elif tipo == 'salida':
            material.cantidad -= cantidad

        movimiento = MovimientoMaterial(material_id=material_id, tipo=tipo, cantidad=cantidad, trabajador=trabajador)
        db.session.add(movimiento)
        db.session.commit()
        flash('Movimiento registrado exitosamente.')
        return redirect(url_for('material_dashboard'))

    materiales = Material.query.all()
    return render_template('nuevo_movimiento_material.html', materiales=materiales)

# Nuevas rutas para Herramienta
@app.route('/herramienta', methods=['GET'])
def herramienta_dashboard():
    # Obtener la página actual de la solicitud (o por defecto la página 1)
    page = request.args.get('page', 1, type=int)
    per_page = 10  # Cantidad de elementos por página

    # Consultar las herramientas con paginación
    herramientas = Herramienta.query.paginate(page=page, per_page=per_page)

    # Variables de paginación
    total_pages = herramientas.pages
    # Ajustar los valores para que 'start_page' y 'end_page' no salgan de rango
    start_page = max(1, page - 5)  # Mostrar desde 5 páginas antes de la actual
    end_page = min(total_pages, page + 5)  # Mostrar hasta 5 páginas después de la actual

    if total_pages == 0:  # Si no hay páginas, establece los valores predeterminados
        start_page = 1
        end_page = 1

    # Renderizar la plantilla y pasar las variables
    return render_template('herramienta_dashboard.html',
                           herramientas=herramientas,
                           page=page,
                           total_pages=total_pages,
                           start_page=start_page,
                           end_page=end_page)

@app.route('/nuevo_herramienta', methods=['GET', 'POST'])
def nuevo_herramienta():
    if request.method == 'POST':
        # Obtener los datos del formulario
        nombre = request.form['nombre']
        vida_util = request.form['vida_util']
        unidad = request.form['unidad']
        cantidad = int(request.form['cantidad'])
        proyecto = request.form['proyecto']
        area_colocacion = request.form['area_colocacion']
        persona_asignada = request.form['persona_asignada']

        # Verificar si ya existe una herramienta con los mismos datos (excepto la cantidad)
        herramienta_existente = Herramienta.query.filter_by(
            nombre=nombre,
            vida_util=vida_util,
            unidad=unidad,
            proyecto=proyecto,
            area_colocacion=area_colocacion,
            persona_asignada=persona_asignada
        ).first()

        if herramienta_existente:
            # Si la herramienta ya existe, actualizar la cantidad
            herramienta_existente.cantidad += cantidad
            db.session.commit()
            flash(f'Herramienta existente actualizada. Nueva cantidad: {herramienta_existente.cantidad}', 'success')
        else:
            # Si no existe, crear una nueva herramienta
            nueva_herramienta = Herramienta(
                nombre=nombre,
                vida_util=vida_util,
                unidad=unidad,
                cantidad=cantidad,
                proyecto=proyecto,
                area_colocacion=area_colocacion,
                persona_asignada=persona_asignada
            )
            db.session.add(nueva_herramienta)
            db.session.commit()
            flash('Herramienta creada exitosamente', 'success')

        return redirect(url_for('herramienta_dashboard'))

    return render_template('nuevo_herramienta.html')



@app.route('/herramienta/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_herramienta(id):
    if current_user.role != 'admin':
        flash('Acceso denegado')
        return redirect(url_for('herramienta_dashboard'))

    herramienta = Herramienta.query.get_or_404(id)

    if request.method == 'POST':
        herramienta.nombre = request.form['nombre']
        herramienta.vida_util = request.form['vida_util']
        herramienta.unidad = request.form['unidad']
        herramienta.cantidad = request.form['cantidad']
        herramienta.proyecto = request.form['proyecto']  # Nuevo campo
        herramienta.area_colocacion = request.form['area_colocacion']  # Nuevo campo
        herramienta.persona_asignada = request.form['persona_asignada']  # Nuevo campo

        db.session.commit()
        flash('Herramienta actualizada exitosamente.')
        return redirect(url_for('herramienta_dashboard'))

    return render_template('editar_herramienta.html', herramienta=herramienta)

@app.route('/herramienta/eliminar/<int:id>')
@login_required
def eliminar_herramienta(id):
    if current_user.role != 'admin':
        flash('Acceso denegado')
        return redirect(url_for('herramienta_dashboard'))
    
    herramienta = Herramienta.query.get_or_404(id)

    # Verificar si hay movimientos relacionados
    if herramienta.movimientos:
        flash('No se puede eliminar la herramienta porque tiene movimientos asociados.')
        return redirect(url_for('herramienta_dashboard'))

    db.session.delete(herramienta)
    db.session.commit()
    flash('Herramienta eliminada exitosamente.')
    return redirect(url_for('herramienta_dashboard'))


@app.route('/herramienta/movimiento/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo_movimiento_herramienta():
    if current_user.role != 'admin':
        flash('Acceso denegado')
        return redirect(url_for('herramienta_dashboard'))
    
    if request.method == 'POST':
        herramienta_id = request.form.get('herramienta_id')
        tipo = request.form['tipo']
        cantidad = int(request.form['cantidad'])  # Convertir la cantidad a entero
        trabajador = request.form.get('trabajador') if tipo == 'salida' else None
        
        herramienta = Herramienta.query.get(herramienta_id)
        
        # Validar cantidad mayor a 0
        if cantidad <= 0:
            flash('La cantidad debe ser mayor a 0.')
            return redirect(url_for('nuevo_movimiento_herramienta'))
        
        # Validar que la salida no exceda el stock disponible
        if tipo == 'salida' and cantidad > herramienta.cantidad:

            flash(f'No hay suficiente stock. Actualmente hay {herramienta.cantidad} unidades de {herramienta.nombre}.')
            return redirect(url_for('nuevo_movimiento_herramienta'))
        
        movimiento = MovimientoHerramienta(
            herramienta_id=herramienta_id, 
            tipo=tipo, 
            cantidad=cantidad, 
            trabajador=trabajador
        )
        
        # Ajustar la cantidad en stock según el tipo de movimiento
        if tipo == 'entrada':
            herramienta.cantidad += cantidad
        elif tipo == 'salida':
            herramienta.cantidad -= cantidad
        
        db.session.add(movimiento)
        db.session.commit()
        flash('Movimiento registrado exitosamente.')
        return redirect(url_for('herramienta_dashboard'))
    
    herramientas = Herramienta.query.all()
    return render_template('nuevo_movimiento_herramienta.html', herramientas=herramientas)



from flask import request
@app.route('/buscar_herramienta_asignada', methods=['GET'])
@login_required
def buscar_herramienta_asignada():
    # Obtener el término de búsqueda desde la URL
    search_query = request.args.get('search', '')

    # Página actual
    page = request.args.get('page', 1, type=int)
    per_page = 10  # Cantidad de elementos por página

    # Buscar herramientas que coincidan parcialmente con el nombre o la persona asignada
    herramientas = Herramienta.query.filter(
        (Herramienta.nombre.ilike(f'%{search_query}%')) | 
        (Herramienta.persona_asignada.ilike(f'%{search_query}%'))
    ).paginate(page=page, per_page=per_page)

    # Variables de paginación
    total_pages = herramientas.pages
    start_page = max(1, page - 5)  # Mostrar desde 5 páginas antes de la actual
    end_page = min(total_pages, page + 5)  # Mostrar hasta 5 páginas después de la actual

    if total_pages == 0:  # Si no hay páginas, establecer valores predeterminados
        start_page = 1
        end_page = 1

    # Renderizar la plantilla y pasar las variables
    return render_template('herramienta_dashboard.html',
                           herramientas=herramientas,
                           search_query=search_query,
                           page=page,
                           total_pages=total_pages,
                           start_page=start_page,
                           end_page=end_page)





from openpyxl.styles import Alignment
# Ruta del dashboard principal
@app.route('/')
def dashboard():
    return render_template('dashboard.html')
# Ruta para generar requisición (Formulario de requisición)
@app.route('/requisicion', methods=['GET', 'POST'])
def generar_requisicion():
    if request.method == 'POST':
        # Recibir los datos del formulario (varios materiales)
        skus = request.form.getlist('sku[]')
        materiales = request.form.getlist('material[]')
        tipos = request.form.getlist('tipo[]')
        unidades = request.form.getlist('unidad[]')
        cantidades = request.form.getlist('cantidad[]')

        # Cargar la plantilla de Excel
        ruta_plantilla = os.path.join('templates', 'templates_excel', 'req.xlsx')  # Asegúrate de la ruta correcta
        wb = openpyxl.load_workbook(ruta_plantilla)
        hoja = wb.active  # Trabaja en la primera hoja

        # Insertar cada fila de datos en la plantilla (empezar en la fila 10)
        fila_inicial = 10
        for i in range(len(skus)):
            # Escribir SKU en la columna B (B10 en adelante)
            if not isinstance(hoja[f'B{fila_inicial}'], openpyxl.cell.cell.MergedCell):
                hoja[f'B{fila_inicial}'].value = skus[i]

            # Escribir el material en la primera celda de la combinación (columna C)
            if not isinstance(hoja[f'C{fila_inicial}'], openpyxl.cell.cell.MergedCell):
                hoja[f'C{fila_inicial}'].value = materiales[i]

            # Colocar tipo en la columna F
            if not isinstance(hoja[f'F{fila_inicial}'], openpyxl.cell.cell.MergedCell):
                hoja[f'F{fila_inicial}'].value = tipos[i]

            # Colocar unidad en la columna G
            if not isinstance(hoja[f'G{fila_inicial}'], openpyxl.cell.cell.MergedCell):
                hoja[f'G{fila_inicial}'].value = unidades[i]

            # Colocar cantidad en la columna H
            if not isinstance(hoja[f'H{fila_inicial}'], openpyxl.cell.cell.MergedCell):
                hoja[f'H{fila_inicial}'].value = cantidades[i]

            fila_inicial += 1

        # Guardar el archivo Excel modificado en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Enviar el archivo como descarga
        return send_file(output, as_attachment=True, download_name="requisicion_completada.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # Si es una solicitud GET, mostrar el formulario de requisición
    return render_template('requisicion_form.html')












from copy import copy
import copy  # Asegúrate de importar el módulo copy
from openpyxl.utils import range_boundaries

from openpyxl.styles import Alignment, Border, Side




# Importar la función copy para copiar propiedades de celdas
from copy import copy

# Importar la función copy para copiar propiedades de celdas
from copy import copy


# Importar la función copy para copiar propiedades de celdas
from copy import copy

@app.route('/nueva_orden_compra', methods=['GET', 'POST'])
@login_required
def nueva_orden_compra():
    if request.method == 'POST':
        # Procesar los datos del formulario
        numero = request.form['numero']
        fecha = request.form['fecha']
        contacto = request.form['contacto']
        rfc = request.form['rfc']
        condicion = request.form['condicion']
        productos = request.form.getlist('productos[]')
        cantidades = request.form.getlist('cantidades[]')
        precios = request.form.getlist('precios[]')
        
        # Cargar la plantilla de Excel
        ruta_plantilla = 'templates/templates_excel/Ocompra.xlsx'
        wb = openpyxl.load_workbook(ruta_plantilla)
        hoja = wb.active

        # Inserta los datos de la cabecera de la orden de compra
        hoja['I4'] = numero
        hoja['I5'] = fecha
        hoja['C12'] = contacto
        hoja['F12'] = rfc
        hoja['I12'] = condicion

        # Define la alineación centrada
        alineacion_centrada = Alignment(horizontal="center", vertical="center")

        # Inserta los productos, cantidades y precios en las filas correspondientes
        fila_inicial = 15  # Ajusta según la estructura de tu plantilla
        for i, (producto, cantidad, precio) in enumerate(zip(productos, cantidades, precios)):
            hoja.unmerge_cells(f'C{fila_inicial}:H{fila_inicial}')  # Producto abarca de C a H

            # Inserta los valores en las celdas correspondientes
            hoja[f'B{fila_inicial}'] = cantidad  # Columna B para la cantidad
            hoja[f'C{fila_inicial}'] = producto  # Columna C, la descripción del producto
            hoja[f'I{fila_inicial}'] = precio    # Columna I para el precio unitario
            hoja[f'K{fila_inicial}'] = f"=B{fila_inicial}*I{fila_inicial}"  # Fórmula para calcular el importe (cantidad * precio unitario)

            # Combinar de nuevo las celdas para Producto (C-H)
            hoja.merge_cells(f'C{fila_inicial}:H{fila_inicial}')

            # Aplicar la alineación centrada a las celdas
            hoja[f'B{fila_inicial}'].alignment = alineacion_centrada  # Centrar la cantidad
            hoja[f'C{fila_inicial}'].alignment = alineacion_centrada  # Centrar la descripción del producto (combinado)
            hoja[f'I{fila_inicial}'].alignment = alineacion_centrada  # Centrar el precio unitario
            hoja[f'K{fila_inicial}'].alignment = alineacion_centrada  # Centrar el importe total

            fila_inicial += 1

        # Cálculo del Subtotal
        fila_subtotal = 36  # Ajustar según la ubicación de tu celda de Subtotal
        hoja[f'K{fila_subtotal}'] = f"=SUM(K15:K{fila_inicial-1})"  # Sumar todos los importes de los productos

        # Cálculo del IVA (16%)
        fila_iva = fila_subtotal + 1  # La celda justo debajo del Subtotal
        hoja[f'K{fila_iva}'] = f"=K{fila_subtotal}*0.16"  # Calcula el 16% del subtotal

        # Cálculo del Total (Subtotal + IVA)
        fila_total = fila_iva + 1  # La celda justo debajo del IVA
        hoja[f'K{fila_total}'] = f"=K{fila_subtotal}+K{fila_iva}"  # Sumar Subtotal e IVA para obtener el total

        # Guarda el archivo en memoria para enviarlo al usuario
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(output, as_attachment=True, download_name="orden_compra.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # Si es una solicitud GET, mostramos el formulario de la Orden de Compra
    return render_template('nueva_orden_compra.html')











































# Configuración de Flask-Mail para Gmail
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USE_SSL'] = False
app.config['MAIL_USERNAME'] = 's.fernando3640@gmail.com'  # Reemplaza con tu correo
app.config['MAIL_PASSWORD'] = 'FerFaFer18051212'  # Reemplaza con tu contraseña

mail = Mail(app)

@app.route('/enviar_requisicion', methods=['POST'])
def enviar_requisicion():
    # Obtener los datos de la requisición desde el formulario
    sku = request.form['sku']
    material = request.form['material']
    tipo = request.form['tipo']
    unidad = request.form['unidad']
    cantidad = request.form['cantidad']

    # Generar el PDF (usando tu método de generación de PDFs)
    pdf_content = generar_pdf_requisicion(sku, material, tipo, unidad, cantidad)

    # Guardar el PDF temporalmente
    pdf_path = 'requisicion.pdf'
    with open(pdf_path, 'wb') as f:
        f.write(pdf_content)

    # Crear y enviar el correo electrónico
    msg = Message(
        'Requisición de Material',
        sender='s.fernando3640@gmail.com',
        recipients=['23fernandoemanuel@gmail.com']  # Reemplaza con el destinatario
    )
    msg.body = f'Se adjunta la requisición del material {material}.'

    with app.open_resource(pdf_path) as pdf:
        msg.attach("requisicion.pdf", "application/pdf", pdf.read())

    mail.send(msg)

    # Eliminar el archivo temporal
    os.remove(pdf_path)

    return 'Correo enviado con éxito'

    # Envía el correo
    mail.send(msg)
        
        # Si todo salió bien, puedes agregar un mensaje de éxito
    flash('Requisición enviada con éxito', 'success')
         

def generar_pdf_requisicion(sku, material, tipo, unidad, cantidad):
    # Aquí iría tu lógica para generar el PDF
    # Por ejemplo, si estás usando pdfkit, algo así:
    rendered = render_template('requisicion_template.html', sku=sku, material=material, tipo=tipo, unidad=unidad, cantidad=cantidad)
    pdf_content = pdfkit.from_string(rendered, False)
    return pdf_content



# Ruta para mostrar el formulario de requisición
@app.route('/nueva_requisicion')
def nueva_requisicion():
    return render_template('nueva_requisicion.html')

# Ruta para crear una nueva requisición
@app.route('/crear_requisicion', methods=['POST'])
def crear_requisicion():
    nombre = request.form['nombre']
    cantidad = request.form['cantidad']
    
    # Guardar la requisición en un archivo CSV
    with open('requisiciones.csv', mode='a', newline='') as file:
        writer = csv.writer(file)
        writer.writerow([nombre, cantidad])
    
    flash('Requisición creada exitosamente.')
    return redirect(url_for('nueva_requisicion'))

# Mostrar las requisiciones en una tabla (si decides mostrar las requisiciones en algún lugar)
@app.route('/ver_requisiciones')
def ver_requisiciones():
    requisiciones = []
    with open('requisiciones.csv', mode='r') as file:
        reader = csv.reader(file)
        for row in reader:
            requisiciones.append(row)
    
    return render_template('ver_requisiciones.html', requisiciones=requisiciones)


@app.route('/obtener_todas_herramientas', methods=['GET'])
def obtener_todas_herramientas():
    search_query = request.args.get('search', None)
    
    # Si hay un parámetro de búsqueda, aplica el filtro
    if search_query:
        herramientas = Herramienta.query.filter(
            Herramienta.nombre.contains(search_query) | 
            Herramienta.persona_asignada.contains(search_query)
        ).all()
    else:
        # Si no hay búsqueda, devuelve todas las herramientas
        herramientas = Herramienta.query.all()

    # Retorna los datos en formato JSON
    return jsonify([{
        'id': h.id,
        'nombre': h.nombre,
        'vida_util': h.vida_util,
        'unidad': h.unidad,
        'cantidad': h.cantidad,
        'proyecto': h.proyecto,
        'area_colocacion': h.area_colocacion,
        'persona_asignada': h.persona_asignada
    } for h in herramientas])




@app.route('/entrada_herramienta/<int:herramienta_id>', methods=['POST'])
def entrada_herramienta(herramienta_id):
    herramienta = Herramienta.query.get_or_404(herramienta_id)
    
    try:
        # Obtener los datos del formulario
        cantidad_entrada = int(request.form['cantidad_entrada'])
        proyecto_asignado = request.form['proyecto_asignado']
        area_colocacion = request.form.get('area_colocacion')  # Nuevo campo de área de colocación
        persona_asignada = request.form.get('persona_asignada')  # Nuevo campo de persona asignada

        if cantidad_entrada <= 0:
            flash("La cantidad ingresada debe ser mayor que cero", "error")
            return redirect(url_for('herramienta_dashboard'))

        # Verificar si ya existe un registro de inventario con todos los campos iguales
        inventario = Inventario.query.filter_by(
            herramienta_id=herramienta_id, 
            proyecto=proyecto_asignado, 
            area_colocacion=area_colocacion, 
            persona_asignada=persona_asignada
        ).first()

        if inventario:
            # Si ya existe un inventario para este proyecto y todos los campos son iguales, actualiza el stock
            inventario.cantidad += cantidad_entrada
            flash(f'Se han agregado {cantidad_entrada} unidades al proyecto {proyecto_asignado}. Stock actualizado.', 'success')
        else:
            # Si no existe, crea un nuevo registro en el inventario
            nuevo_inventario = Inventario(
                herramienta_id=herramienta_id,
                proyecto=proyecto_asignado,
                cantidad=cantidad_entrada,
                area_colocacion=area_colocacion,
                persona_asignada=persona_asignada
            )
            db.session.add(nuevo_inventario)
            flash(f'Se han creado {cantidad_entrada} unidades para el proyecto {proyecto_asignado}', 'success')

        db.session.commit()

    except Exception as e:
        flash(f'Error al registrar la entrada: {str(e)}', 'error')

    return redirect(url_for('herramienta_dashboard'))


@app.route('/salida_herramienta/<int:herramienta_id>', methods=['POST'])
def salida_herramienta(herramienta_id):
    # Obtener la herramienta por su ID
    herramienta = Herramienta.query.get_or_404(herramienta_id)
    
    try:
        # Obtener la cantidad que se está retirando desde el formulario
        cantidad_salida = int(request.form['cantidad_salida'])
        persona_asignada = request.form['persona_asignada']  # A quién se asigna
        proyecto_asignado = request.form['proyecto_asignado']  # Proyecto asignado

        if cantidad_salida <= 0:
            flash("La cantidad retirada debe ser mayor que cero", "error")
            return redirect(url_for('herramienta_dashboard'))

        if cantidad_salida > herramienta.cantidad:
            flash(f'No hay suficientes unidades de {herramienta.nombre} en el inventario', 'error')
            return redirect(url_for('herramienta_dashboard'))

        # Actualizar la cantidad en el inventario
        herramienta.cantidad -= cantidad_salida
        herramienta.persona_asignada = persona_asignada
        herramienta.proyecto = proyecto_asignado
        db.session.commit()

        flash(f'Se han asignado {cantidad_salida} unidades de {herramienta.nombre} a {persona_asignada} para el proyecto {proyecto_asignado}', 'success')
    except Exception as e:
        flash('Error al asignar la herramienta: ' + str(e), 'error')

    return redirect(url_for('herramienta_dashboard'))










































if __name__ == '__main__':
    app.run(debug=True)
